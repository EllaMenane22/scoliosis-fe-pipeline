r"""
Final stage of the pipeline. Opens each patient's solved Marc .t16, measures the
Cobb angle at the first and last increment (so before and after the brace load is
applied), and lines the model's predicted correction up against the clinical
before/in-brace correction. Everything goes into one colour-coded Excel table.

Has to be run with the Marc-bundled Python, otherwise py_post won't import.

Output:
...\Automated models\Updated automated\cobb_results_updated_automated_formatted.xlsx
"""

import os
import sys
import csv
import math
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# Paths

BASE_DIR = Path(
    r"C:\Users\<username>\OneDrive - Imperial College London\Year 4\FYP"
)

OUTPUT_ROOT = Path(
    r"C:\Users\<username>\OneDrive - Imperial College London\Year 4\FYP\Patient_1_final\Automated models\Updated automated"
)

CLINICAL_CSV = BASE_DIR / "clinical_patient_data.csv"

RESULTS_XLSX = OUTPUT_ROOT / "cobb_results_updated_automated_formatted.xlsx"

# Where py_post and its DLLs live. These get added to the path at startup so the
# Marc post-processing library can be imported. They're specific to this Marc
# install, so they'd need editing on another machine.
MARC_PYTHON_PATHS = [
    r"C:\Users\<username>\OneDrive - Imperial College London\Year 4\FYP\Patient_1_final\shlib\win64",
    r"C:\Program Files\MSC.Software\Marc\2024.1.0\mentat2024.1\shlib\win64",
    r"C:\Program Files\MSC.Software\Marc\2024.1.0\mentat2024.1\python\WIN8664",
    r"C:\Program Files\MSC.Software\Marc\2024.1.0\mentat2024.1\python\WIN8664\Lib",
]


# Clinical CSV column names, gathered here so a header change is a one-line edit

PATIENT_ID_COL = "Patient's Number (New System)"

CLINICAL_BEFORE_COLS = {
    "Thoracic": "Thoracic Cobb Angle (Before Brace)",
    "Lumbar": "Lumbar Cobb Angle (Before Brace)",
    "Thoracolumbar": "Thoraco Lumbar Cobb Angle (Before Brace)",
}

# In-brace angle headers, with the different spellings the spreadsheet has used.
CLINICAL_INBRACE_COL_OPTIONS = {
    "Thoracic": [
        "Thoracic Cobb Angle (In Brace)",
        "Thoracic Cobb Angle (In-Brace)",
        "Thoracic Cobb Angle (After Brace)",
    ],
    "Lumbar": [
        "Lumbar Cobb Angle (In Brace)",
        "Lumbar Cobb Angle (In-Brace)",
        "Lumbar Cobb Angle (After Brace)",
    ],
    "Thoracolumbar": [
        "Thoraco Lumbar Cobb Angle (In Brace)",
        "Thoraco Lumbar Cobb Angle (In-Brace)",
        "Thoraco Lumbar Cobb Angle (After Brace)",
    ],
}


# Spine levels, bottom to top. Has to match the order used in coords.py or the
# node-to-vertebra mapping below would be wrong.

SPINE_LEVELS = [
    "L5", "L4", "L3", "L2", "L1",
    "T12", "T11", "T10", "T9", "T8", "T7", "T6",
    "T5", "T4", "T3", "T2", "T1",
    "C7", "C6", "C5", "C4", "C3", "C2", "C1",
]


# Getting py_post to import

def setup_py_post_paths() -> None:
    for folder in MARC_PYTHON_PATHS:
        if os.path.exists(folder) and folder not in sys.path:
            sys.path.append(folder)

    # On newer Windows the DLL folders also have to be registered explicitly.
    if hasattr(os, "add_dll_directory"):
        for folder in MARC_PYTHON_PATHS:
            if os.path.exists(folder):
                try:
                    os.add_dll_directory(folder)
                except OSError:
                    pass

    # Fail loudly here rather than halfway through, since nothing works without it.
    try:
        import py_post  # noqa: F401
        print("py_post import check: PASSED")
    except Exception as exc:
        print("\nERROR: Could not import py_post.")
        print(f"Python executable: {sys.executable}")
        print("Original error:")
        print(exc)
        raise


# Small vector helpers used to measure the Cobb angle. The Cobb angle is the
# angle between the end-plate lines of the two end vertebrae, so each vertebra's
# axis is turned into its perpendicular (the end-plate direction) and the angle
# between those two perpendiculars is taken.

def vec(a, b):
    return (b[0] - a[0], b[1] - a[1])


def perp(v):
    return (-v[1], v[0])


def norm(v):
    return math.hypot(v[0], v[1])


def angle_between_lines(u, v):
    nu, nv = norm(u), norm(v)

    if nu == 0.0 or nv == 0.0:
        raise ValueError("Zero-length vector - check vertebra labels.")

    dot = abs(u[0] * v[0] + u[1] * v[1])
    c = max(-1.0, min(1.0, dot / (nu * nv)))

    return math.degrees(math.acos(c))


# Reading the per-patient CSVs and the clinical sheet

def clean_float(value):
    if value is None:
        return None

    s = str(value).strip()

    if s == "":
        return None

    try:
        return float(s)
    except ValueError:
        return None


def patient_label_from_row(value):
    s = str(value).strip()
    m = re.search(r"\d+", s)

    if not m:
        return None

    n = int(m.group())
    return f"P{n:02d}" if n < 100 else f"P{n}"


def patient_sort_key(folder: Path):
    m = re.search(r"\d+", folder.name)
    return int(m.group()) if m else 999999


def read_metadata(folder: Path, pid: str):
    path = folder / f"{pid}_model_metadata.csv"

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        raise ValueError(f"Metadata file is empty: {path}")

    row = {k.strip(): (v or "").strip() for k, v in rows[0].items()}

    return {
        "region": row["TargetCurveRegion"],
        "convexity": row["Convexity"],
        "lower_end": row["LowerModelCurveEnd"].upper(),
        "upper_end": row["UpperModelCurveEnd"].upper(),
    }


def build_vertebra_map(folder: Path, pid: str):
    path = folder / f"{pid}_spine.csv"

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    n_nodes = len(rows)
    vmap = {}

    # Each level owns two nodes (bottom 2i+1, top 2i+2), matching coords.py.
    for i, label in enumerate(SPINE_LEVELS):
        bot = 2 * i + 1
        top = 2 * i + 2

        if top <= n_nodes:
            vmap[label] = {"bot": bot, "top": top}

    return vmap


def read_clinical(csv_path: Path):
    out = {}

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = [h.strip() for h in (reader.fieldnames or [])]

        for raw in reader:
            row = {k.strip(): v for k, v in raw.items() if k is not None}
            pid = patient_label_from_row(row.get(PATIENT_ID_COL, ""))

            if not pid:
                continue

            for region, before_col in CLINICAL_BEFORE_COLS.items():
                before = clean_float(row.get(before_col))

                if before is None:
                    continue

                inbrace = None

                for opt in CLINICAL_INBRACE_COL_OPTIONS[region]:
                    if opt in fieldnames:
                        inbrace = clean_float(row.get(opt))

                        if inbrace is not None:
                            break

                out[pid] = {
                    "region": region,
                    "before": before,
                    "inbrace": inbrace,
                }

                break

    return out


# Measuring the Cobb angle from the .t16

def deformed_position(p, node_id, inc):
    p.moveto(inc)

    idx = p.node_sequence(node_id)

    if idx < 0:
        raise ValueError(f"Node {node_id} not found in post file.")

    n = p.node(idx)

    try:
        dx, dy, dz = p.node_displacement(idx)
    except Exception:
        dx, dy, dz = 0.0, 0.0, 0.0

    return (n.x + dx, n.y + dy)


def vertebra_endplate(p, vname, inc, vmap):
    info = vmap[vname]

    top_pt = deformed_position(p, info["top"], inc)
    bot_pt = deformed_position(p, info["bot"], inc)

    axis = vec(bot_pt, top_pt)

    if norm(axis) == 0.0:
        raise ValueError(f"Vertebra {vname} has zero-length axis.")

    return perp(axis)


def measure_cobb(p, top_name, bot_name, inc, vmap):
    return angle_between_lines(
        vertebra_endplate(p, top_name, inc, vmap),
        vertebra_endplate(p, bot_name, inc, vmap),
    )


def measure_model_cobb(t16_path, vmap, top_name, bot_name):
    from py_post import post_open

    p = post_open(str(t16_path))
    ninc = p.increments()

    if ninc < 2:
        p.close()
        raise ValueError(f"Only {ninc} increment(s) - job may not have converged.")

    # Increment 0 is the undeformed (initial) shape, the last increment is the
    # fully loaded (in-brace) shape.
    initial = measure_cobb(p, top_name, bot_name, 0, vmap)
    final = measure_cobb(p, top_name, bot_name, ninc - 1, vmap)

    p.close()

    return initial, final


# Turning angles into correction figures

def correction_deg(initial, final):
    if initial is None or final is None:
        return None

    return round(final - initial, 2)


def correction_pct(initial, final):
    # Positive means the curve got smaller, i.e. a real correction. Guards
    # against a zero initial angle so it never divides by zero.
    if initial is None or final is None or initial == 0:
        return None

    return round((initial - final) / initial, 4)


def region_label(region, convexity):
    pretty = {
        "Thoracic": "Thoracic",
        "Lumbar": "Lumbar",
        "Thoracolumbar": "Thoraco-Lumbar",
    }.get(region, region)

    return f"{pretty} ({convexity})" if convexity else pretty


# Building the formatted comparison spreadsheet

def write_formatted_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cobb Angle Comparison"

    NAVY = "1F3B66"
    DARK_GREY = "3F3F3F"
    MODEL_BLUE = "2F5597"
    CLINICAL_GREEN = "548235"
    DELTA_BROWN = "9E480E"

    LIGHT_GREY = "E7E6E6"
    LIGHT_BLUE = "DDEBF7"
    LIGHT_GREEN = "E2F0D9"
    LIGHT_ORANGE = "FCE4D6"
    LIGHT_YELLOW = "FFF2CC"

    WHITE = "FFFFFF"

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:M1")
    title = ws["A1"]
    title.value = "Cobb Angle Comparison — Model Prediction vs Clinical Data"
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.font = Font(color=WHITE, bold=True, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:D2")
    ws.merge_cells("E2:H2")
    ws.merge_cells("I2:L2")
    ws.merge_cells("M2:M3")

    group_headers = {
        "A2": ("Patient", DARK_GREY),
        "B2": ("Curve Region\n(Convexity)", DARK_GREY),
        "C2": ("Vertebral Levels Used", DARK_GREY),
        "E2": ("Model (Predicted)", MODEL_BLUE),
        "I2": ("Clinical (Actual)", CLINICAL_GREEN),
        "M2": ("Correction Δ\n(Model − Clinical)", DELTA_BROWN),
    }

    for cell_ref, (text, colour) in group_headers.items():
        cell = ws[cell_ref]
        cell.value = text
        cell.fill = PatternFill("solid", fgColor=colour)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    subheaders = [
        None,
        None,
        "Upper End\nVertebra",
        "Lower End\nVertebra",
        "Initial Cobb (°)",
        "Final Cobb (°)",
        "Correction (°)",
        "Correction (%)",
        "Initial Cobb (°)\n(Before Brace)",
        "Final Cobb (°)\n(In-Brace)",
        "Correction (°)",
        "Correction (%)",
        None,
    ]

    for col_idx, header in enumerate(subheaders, start=1):
        cell = ws.cell(row=3, column=col_idx)

        if header:
            cell.value = header

        if col_idx in [1, 2, 3, 4]:
            fill = DARK_GREY
        elif 5 <= col_idx <= 8:
            fill = MODEL_BLUE
        elif 9 <= col_idx <= 12:
            fill = CLINICAL_GREEN
        else:
            fill = DELTA_BROWN

        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=WHITE, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 40

    start_row = 4

    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if c_idx in [1, 2]:
                fill = LIGHT_GREY
            elif c_idx in [3, 4]:
                fill = LIGHT_YELLOW
            elif 5 <= c_idx <= 8:
                fill = LIGHT_BLUE
            elif 9 <= c_idx <= 12:
                fill = LIGHT_GREEN
            elif c_idx == 13:
                fill = LIGHT_ORANGE
            else:
                fill = "FFFFFF"

            cell.fill = PatternFill("solid", fgColor=fill)

            if c_idx == 1:
                cell.font = Font(bold=True)

            if c_idx in [5, 6, 7, 9, 10, 11]:
                cell.number_format = '0.0°'

            if c_idx in [8, 12]:
                cell.number_format = '0.0%'

            if c_idx == 13:
                cell.number_format = '+0.0%;-0.0%;0.0%'
                cell.font = Font(bold=True)

    widths = {
        "A": 10,
        "B": 22,
        "C": 13,
        "D": 13,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 15,
        "J": 15,
        "K": 13,
        "L": 13,
        "M": 18,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row_idx in range(start_row, start_row + len(rows)):
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:M{start_row + len(rows) - 1}"

    for row in ws.iter_rows(
        min_row=1,
        max_row=start_row + len(rows) - 1,
        min_col=1,
        max_col=13,
    ):
        for cell in row:
            cell.border = border

    RESULTS_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RESULTS_XLSX)


# Main: loop patients, measure each .t16, write the table

def main():
    setup_py_post_paths()

    print("\n" + "=" * 72)
    print("COBB RESULTS TABLE GENERATOR — FORMATTED XLSX")
    print("=" * 72)
    print("Patient result folder:")
    print(f"  {OUTPUT_ROOT}")
    print("Clinical CSV:")
    print(f"  {CLINICAL_CSV}")
    print("Results XLSX:")
    print(f"  {RESULTS_XLSX}")
    print("=" * 72 + "\n")

    if not OUTPUT_ROOT.exists():
        raise FileNotFoundError(f"OUTPUT_ROOT not found:\n{OUTPUT_ROOT}")

    if not CLINICAL_CSV.exists():
        raise FileNotFoundError(f"Clinical CSV not found:\n{CLINICAL_CSV}")

    clinical = read_clinical(CLINICAL_CSV)

    patient_dirs = sorted(
        [d for d in OUTPUT_ROOT.iterdir() if d.is_dir()],
        key=patient_sort_key,
    )

    if not patient_dirs:
        print(f"No patient folders found in:\n{OUTPUT_ROOT}")
        return

    rows = []
    done = 0
    pending = 0
    skipped = 0

    for folder in patient_dirs:
        pid = folder.name

        try:
            meta = read_metadata(folder, pid)
        except Exception as e:
            print(f"  [{pid}] no metadata ({e}) - skipping")
            skipped += 1
            continue

        try:
            vmap = build_vertebra_map(folder, pid)
        except Exception as e:
            print(f"  [{pid}] spine CSV read error ({e}) - skipping")
            skipped += 1
            continue

        top_name = meta["upper_end"]
        bot_name = meta["lower_end"]

        if top_name not in vmap or bot_name not in vmap:
            print(
                f"  [{pid}] end vertebra not found in model map "
                f"(top={top_name}, bottom={bot_name}) - skipping"
            )
            skipped += 1
            continue

        t16 = folder / f"{pid}.t16"

        m_init = None
        m_final = None

        if t16.exists():
            try:
                m_init, m_final = measure_model_cobb(
                    t16,
                    vmap,
                    top_name,
                    bot_name,
                )
                m_init = round(m_init, 2)
                m_final = round(m_final, 2)
                done += 1
            except Exception as e:
                print(f"  [{pid}] t16 read error: {e}")
        else:
            pending += 1
            print(f"  [{pid}] .t16 not found yet - row left blank for model")

        cl = clinical.get(pid, {})
        c_before = cl.get("before")
        c_inbrace = cl.get("inbrace")

        m_corr_pct = correction_pct(m_init, m_final)
        c_corr_pct = correction_pct(c_before, c_inbrace)

        delta = (
            round(m_corr_pct - c_corr_pct, 4)
            if (m_corr_pct is not None and c_corr_pct is not None)
            else None
        )

        rows.append([
            pid,
            region_label(meta["region"], meta["convexity"]),
            top_name,
            bot_name,
            m_init,
            m_final,
            correction_deg(m_init, m_final),
            m_corr_pct,
            c_before,
            c_inbrace,
            correction_deg(c_before, c_inbrace),
            c_corr_pct,
            delta,
        ])

    write_formatted_excel(rows)

    print("\n" + "=" * 72)
    print("Results written:")
    print(f"  {RESULTS_XLSX}")
    print(f"Patients in table: {len(rows)}")
    print(f"Model results in : {done}")
    print(f"Still pending    : {pending}  (.t16 not generated yet)")
    print(f"Skipped          : {skipped}")
    print("=" * 72)


if __name__ == "__main__":
    main()
